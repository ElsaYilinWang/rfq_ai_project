"""
RFQ Parser
===========
Parses an RFQ Excel spreadsheet (.xlsm) and produces:
1. A normalized JSON output file
2. A modified Excel file with comments on Material cells
"""

import json
import re
from pathlib import Path
from dataclasses import asdict

import openpyxl
from openpyxl.comments import Comment

from .schemas import (
    RFQMetadata,
    SourcingIdentifier,
    ExtractedReference,
    LineItem,
    ParsedRFQ
)
from .validators import (
    validate_hard_errors,
    validate_sourcing_identifiers,
    generate_overall_flags
)


class RFQParser:

    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.workbook = None
        self.sheet = None

    # ---------------------------------------------------------------
    # Main entry point
    # ---------------------------------------------------------------

    def parse(self, output_json_folder: str) -> ParsedRFQ:
        """
        Main entry point. Orchestrates the full parsing pipeline.
        Returns a ParsedRFQ object and writes JSON and Excel outputs.
        """
        # Step 1 — Validate file
        self._validate_file()

        # Step 2 — Load workbook
        self.workbook = openpyxl.load_workbook(
            self.file_path, keep_vba=True
        )
        try:
            self.sheet = self.workbook['SPREADSHEET']
        except KeyError:
            raise ValueError(
                f"Hard error: 'SPREADSHEET' sheet not found in file: "
                f"{self.file_path.name}. "
                f"Available sheets: {self.workbook.sheetnames}"
            )

        # Step 3 — Extract metadata
        metadata = self._extract_metadata()

        # Step 4 — Parse line items
        items = self._parse_line_items()

        # Step 5 — Validate hard errors
        validate_hard_errors(metadata.rfq_number, len(items))

        # Step 6 — Validate each item and generate flags
        for item in items:
            item.flags = validate_sourcing_identifiers(item)

        # Step 7 — Generate overall flags
        overall_flags = generate_overall_flags(items)

        # Step 8 — Build ParsedRFQ object
        parsed_rfq = ParsedRFQ(
            metadata=metadata,
            items=items,
            overall_flags=overall_flags
        )

        # Step 9 — Write outputs
        self._write_json_output(parsed_rfq, output_json_folder)
        self._write_excel_comments(parsed_rfq)

        return parsed_rfq

    # ---------------------------------------------------------------
    # Step 1 — Validate file
    # ---------------------------------------------------------------

    def _validate_file(self) -> None:
        """
        Validates that the file exists, is an xlsm file,
        and contains a SPREADSHEET sheet.
        """
        if not self.file_path.exists():
            raise FileNotFoundError(
                f"File not found: {self.file_path}"
            )

        if self.file_path.suffix.lower() not in ['.xlsm', '.xlsx']:
            raise ValueError(
                f"Invalid file type: {self.file_path.suffix}. "
                f"Expected .xlsm or .xlsx"
            )

    # ---------------------------------------------------------------
    # Step 2 — Extract metadata
    # ---------------------------------------------------------------

    def _extract_metadata(self) -> RFQMetadata:
        """
        Extracts RFQ-level metadata from Row 1 and the filename.
        Also validates that the RFQ number in the filename matches
        cell A1 — raises ValueError if they don't match.
        """
        # Extract RFQ number from cell A1
        rfq_number = str(self.sheet.cell(row=1, column=1).value or "").strip()

        # Extract client contact from cell C1
        client_contact = str(self.sheet.cell(row=1, column=3).value or "").strip()

        # Extract date from cell D1
        date = str(self.sheet.cell(row=1, column=4).value or "").strip()

        # Extract internal reference and confirm RFQ number from filename
        # Filename format: 3435.0 6000186511.xlsm
        # or:              3424.0 6000186454 Emerson.xlsm
        filename = self.file_path.stem  # e.g. 3435.0 6000186511

        if ".0 " in filename:
            parts = filename.split(".0 ", 1)
            internal_reference = parts[0].strip()
            filename_rfq_number = parts[1].strip()
        else:
            # Fallback for underscore format (test files)
            parts = filename.split("_")
            internal_reference = parts[0] if parts else ""
            filename_rfq_number = parts[-1] if parts else ""

        # Hard check — filename RFQ number must match cell A1
        if filename_rfq_number != rfq_number:
            raise ValueError(
                f"Hard error: RFQ number mismatch. "
                f"Filename has '{filename_rfq_number}' "
                f"but cell A1 has '{rfq_number}'"
            )

        return RFQMetadata(
            source_file_path=str(self.file_path),
            internal_reference=internal_reference,
            rfq_number=rfq_number,
            client_contact=client_contact,
            date=date
        )

    # ---------------------------------------------------------------
    # Step 3 — Parse all line items
    # ---------------------------------------------------------------

    def _parse_line_items(self) -> list:
        """
        Reads all rows from Row 3 onwards and parses each into
        a LineItem object.
        """
        items = []

        for row in self.sheet.iter_rows(min_row=3, values_only=True):
            # Skip empty rows
            if not any(row):
                continue

            # Skip summary rows (no material number)
            if row[1] is None:
                continue

            item = self._parse_single_item(row)
            items.append(item)

        return items

    # ---------------------------------------------------------------
    # Step 4 — Parse a single line item
    # ---------------------------------------------------------------

    def _parse_single_item(self, row: tuple) -> LineItem:
        """
        Parses one row into a LineItem object.
        Column order based on SPREADSHEET sheet:
        Col 0: line number
        Col 1: Material number
        Col 2: Long description
        Col 3: UOM
        Col 4: Quantity
        Col 5: PN/MODEL/MFR
        Col 6: Lead time Requested
        """
        material_number = str(row[1] or "").strip()
        long_description = str(row[2] or "").strip()
        uom = str(row[3] or "").strip()

        # Parse quantity as integer
        try:
            quantity = int(row[4]) if row[4] is not None else 0
        except (ValueError, TypeError):
            quantity = 0

        # Parse PN/MODEL/MFR
        pn_model_mfr = str(row[5] or "").strip()
        sourcing_identifiers, extracted_refs, pn_flags = self._parse_pn_model_mfr(pn_model_mfr)

        # Parse lead time
        lead_time_str = str(row[6] or "").strip()
        lead_time_date, lead_time_weeks = self._parse_lead_time(lead_time_str)

        return LineItem(
            material_number=material_number,
            long_description=long_description,
            uom=uom,
            quantity=quantity,
            lead_time_date=lead_time_date,
            lead_time_weeks=lead_time_weeks,
            sourcing_identifiers=sourcing_identifiers,
            flags=pn_flags,
            extracted_references=extracted_refs
        )

    # ---------------------------------------------------------------
    # Step 5 — Parse PN/MODEL/MFR
    # ---------------------------------------------------------------

    def _parse_pn_model_mfr(self, pn_model_mfr: str) -> tuple:
        """
        Parses PN/MODEL/MFR field into a list of SourcingIdentifier objects
        and a list of unstructured entries for manual review.

        Returns: (sourcing_identifiers, extracted_references, flags)

        Three patterns handled:
        Pattern 1: 'PART_NUMBER - MANUFACTURER' → clean split
        Pattern 2: 'PART_NUMBER / MANUFACTURER' → clean split
        Pattern 3: no separator → store in extracted_references, flag for review
        """
        if not pn_model_mfr or pn_model_mfr.strip() == "":
            return [], [], ["missing_sourcing_identifier"]

        identifiers = []
        extracted = []
        flags = []
        unstructured_count = 0

        entries = pn_model_mfr.split("\n")

        for entry in entries:
            if not entry.strip():
                continue

            # Pattern 1 — try ' - ' separator
            parts = re.split(r'\s+-\s+', entry, maxsplit=1)
            if len(parts) == 2:
                part_number = parts[0].strip() or None
                manufacturer = parts[1].strip() or None
                identifiers.append(SourcingIdentifier(
                    part_number=part_number,
                    manufacturer=manufacturer
                ))
                continue

            # Pattern 2 — try ' / ' separator
            parts = re.split(r'\s+/\s+', entry, maxsplit=1)
            if len(parts) == 2:
                part_number = parts[0].strip() or None
                manufacturer = parts[1].strip() or None
                identifiers.append(SourcingIdentifier(
                    part_number=part_number,
                    manufacturer=manufacturer
                ))
                continue

            # Pattern 3 — no separator found
            # Store first 5 unstructured entries for manual review
            if unstructured_count < 5:
                extracted.append(ExtractedReference(
                    type="UNSTRUCTURED_IDENTIFIER",
                    value=entry.strip()
                ))
                unstructured_count += 1

        # Flag for manual review if any unstructured entries found
        if extracted:
            flags.append("needs_manual_review")

        return identifiers, extracted, flags

    # ---------------------------------------------------------------
    # Step 6 — Parse lead time
    # ---------------------------------------------------------------

    def _parse_lead_time(self, lead_time_str: str) -> tuple:
        """
        Parses lead time string into date and weeks.
        Expected format: '28/05/2026\n11 Weeks'
        """
        if not lead_time_str:
            return None, None

        lead_time_date = None
        lead_time_weeks = None

        parts = lead_time_str.split("\n")

        for part in parts:
            part = part.strip()

            # Match date pattern DD/MM/YYYY
            if re.match(r"\d{1,2}/\d{1,2}/\d{4}", part):
                lead_time_date = part

            # Match weeks pattern
            match = re.search(r"(\d+)\s*[Ww]eeks?", part)
            if match:
                lead_time_weeks = int(match.group(1))

        return lead_time_date, lead_time_weeks

    # ---------------------------------------------------------------
    # Step 7 — Write JSON output
    # ---------------------------------------------------------------

    def _write_json_output(
        self, parsed_rfq: ParsedRFQ, output_folder: str
    ) -> None:
        """
        Writes the parsed RFQ object to a JSON file.
        Saved to the user-specified output folder.
        """
        output_path = Path(output_folder)
        output_path.mkdir(parents=True, exist_ok=True)

        filename = f"parsed_{parsed_rfq.metadata.rfq_number}.json"
        full_path = output_path / filename

        with open(full_path, "w", encoding="utf-8") as f:
            json.dump(asdict(parsed_rfq), f, indent=2, ensure_ascii=False)

        print(f"JSON output written to: {full_path}")

    # ---------------------------------------------------------------
    # Step 8 — Write Excel comments
    # ---------------------------------------------------------------

    def _write_excel_comments(self, parsed_rfq: ParsedRFQ) -> None:
        """
        Adds comments to the Material cell of each line item
        in the original Excel file.
        Each comment contains all flags and extracted references
        for that item.
        Clears any existing comments before writing new ones.
        """
        # First pass — clear ALL existing comments on Material column
        for row_idx in range(3, self.sheet.max_row + 1):
            self.sheet.cell(row=row_idx, column=2).comment = None

        # Second pass — write new comments where needed
        for row_idx, item in enumerate(parsed_rfq.items, start=3):

            # Only add comment if there are flags or extracted references
            if not item.flags and not item.extracted_references:
                continue

            # Build comment text
            comment_lines = []

            if item.flags:
                comment_lines.append("FLAGS:")
                for flag in item.flags:
                    comment_lines.append(f"  - {flag}")

            if item.extracted_references:
                comment_lines.append("REFERENCES:")
                for ref in item.extracted_references:
                    comment_lines.append(f"  - {ref.type}: {ref.value}")

            comment_text = "\n".join(comment_lines)

            # Add comment to Material cell (column B = column 2)
            comment = Comment(comment_text, "RFQ Parser")
            comment.width = 300
            comment.height = 200
            self.sheet.cell(row=row_idx, column=2).comment = comment

        # Save the workbook back to original location
        self.workbook.save(self.file_path)
        print(f"Excel comments written to: {self.file_path}")

        # Second pass — write new comments where needed
        for row_idx, item in enumerate(parsed_rfq.items, start=3):
            
            # DEBUG — remove after fixing
            if item.material_number == '1000090583':
                print(f"DEBUG flags: {item.flags}")
                print(f"DEBUG extracted_references: {item.extracted_references}")

            # Only add comment if there are flags or extracted references
            if not item.flags and not item.extracted_references:
                continue